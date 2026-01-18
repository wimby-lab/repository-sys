import os
from collections import defaultdict
from urllib.parse import quote

from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import FileResponse, Http404
from django.views.decorators.clickjacking import xframe_options_sameorigin
from django.db.models import Q, Count
from django.utils import timezone
from .models import Document, DocumentFolder
from .forms import (
    DocumentUploadForm,
    DocumentUpdateForm,
    DocumentSearchForm,
    DocumentFolderForm,
)
from .permissions import can_access_document, get_accessible_documents, can_manage_folders
from accounts.utils import log_audit
from accounts.decorators import manager_or_admin_required


PREVIEW_CHAR_LIMIT = 8000
PREVIEW_ROW_LIMIT = 25
PREVIEW_COLUMN_LIMIT = 10
PREVIEW_CELL_LIMIT = 200
PREVIEW_MAX_FILE_SIZE = 5 * 1024 * 1024


class PreviewError(Exception):
    """Raised when a preview cannot be generated."""


def _is_safe_media_path(file_path):
    media_root = os.path.abspath(settings.MEDIA_ROOT)
    try:
        return os.path.commonpath([os.path.abspath(file_path), media_root]) == media_root
    except ValueError:
        return False


def _truncate_text(text, limit=PREVIEW_CHAR_LIMIT):
    if not text:
        return '', False
    truncated = len(text) > limit
    return text[:limit], truncated


def _load_text_preview(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8', errors='replace') as file:
            content = file.read(PREVIEW_CHAR_LIMIT + 1)
    except (OSError, UnicodeError) as exc:
        raise PreviewError('Unable to read text preview.') from exc
    return _truncate_text(content)


def _load_docx_preview(file_path):
    from docx import Document as DocxDocument
    from docx.opc.exceptions import PackageNotFoundError

    try:
        doc = DocxDocument(file_path)
    except (PackageNotFoundError, OSError, ValueError) as exc:
        raise PreviewError('Unable to read Word document preview.') from exc
    content = '\n'.join(
        paragraph.text for paragraph in doc.paragraphs if paragraph.text
    )
    return _truncate_text(content)


def _load_spreadsheet_preview(file_path):
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    workbook = None
    sheet_title = ''
    rows = []
    truncated = False
    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        sheet = workbook.active
        for row_index, row in enumerate(sheet.iter_rows(
            values_only=True,
            max_row=PREVIEW_ROW_LIMIT + 1,
            max_col=PREVIEW_COLUMN_LIMIT + 1,
        )):
            if row_index >= PREVIEW_ROW_LIMIT:
                truncated = True
                break
            row_values = []
            if len(row) > PREVIEW_COLUMN_LIMIT:
                truncated = True
            for cell in row[:PREVIEW_COLUMN_LIMIT]:
                if cell is None:
                    row_values.append('')
                    continue
                cell_value = str(cell)
                if len(cell_value) > PREVIEW_CELL_LIMIT:
                    truncated = True
                    cell_value = f'{cell_value[:PREVIEW_CELL_LIMIT]}…'
                row_values.append(cell_value)
            rows.append(row_values)
        sheet_title = sheet.title
    except (InvalidFileException, OSError, ValueError) as exc:
        raise PreviewError('Unable to read spreadsheet preview.') from exc
    finally:
        if workbook is not None:
            workbook.close()

    return sheet_title, rows, truncated


@login_required
def document_list(request):
    """List documents with search and filter"""
    form = DocumentSearchForm(request.GET)
    
    # Get base queryset with access control - exclude archived documents
    documents = Document.objects.select_related('owner').filter(
        get_accessible_documents(request.user),
        is_archived=False
    ).distinct()
    
    # Apply search filters
    if form.is_valid():
        query = form.cleaned_data.get('query')
        if query:
            documents = documents.filter(
                Q(title__icontains=query) | 
                Q(description__icontains=query) |
                Q(file__icontains=query)
            )
        
        classification = form.cleaned_data.get('classification')
        if classification:
            documents = documents.filter(classification=classification)

        section = form.cleaned_data.get('section')
        if section:
            documents = documents.filter(section=section)
        
        category = form.cleaned_data.get('category')
        if category:
            documents = documents.filter(category__icontains=category)
        
        owner = form.cleaned_data.get('owner')
        if owner:
            documents = documents.filter(owner__username__icontains=owner)
        
        date_from = form.cleaned_data.get('date_from')
        if date_from:
            documents = documents.filter(created_at__gte=date_from)
        
        date_to = form.cleaned_data.get('date_to')
        if date_to:
            documents = documents.filter(created_at__lte=date_to)

    documents = documents.order_by('section', '-created_at')
    documents_count = documents.count()
    folder_map = {folder.key: folder for folder in DocumentFolder.objects.order_by('name')}
    section_counts = {
        entry['section']: entry['total']
        for entry in documents.values('section').annotate(total=Count('id'))
    }
    section_labels = dict(Document.SECTION_CHOICES)
    section_map = defaultdict(list)
    for document in documents:
        section_map[document.section].append(document)

    documents_by_section = []
    used_sections = set()
    for section_value, folder in folder_map.items():
        section_documents = section_map.get(section_value, [])
        if section_documents:
            documents_by_section.append({
                'label': folder.name,
                'documents': section_documents
            })
            used_sections.add(section_value)

    for section_value, section_documents in section_map.items():
        if section_value not in used_sections:
            documents_by_section.append({
                'label': section_labels.get(
                    section_value,
                    section_value.replace('_', ' ').title()
                ),
                'documents': section_documents
            })

    folders = [
        {
            'id': folder.id,
            'key': folder.key,
            'name': folder.name,
            'document_count': section_counts.get(folder.key, 0),
        }
        for folder in folder_map.values()
    ]

    return render(request, 'documents/document_list.html', {
        'documents_by_section': documents_by_section,
        'documents_count': documents_count,
        'form': form,
        'folders': folders,
        'can_manage_folders': can_manage_folders(request.user)
    })


@login_required
@manager_or_admin_required
def folder_create(request):
    """Create a new document folder"""
    if request.method == 'POST':
        form = DocumentFolderForm(request.POST)
        if form.is_valid():
            folder = form.save()

            log_audit(
                request.user,
                'DOCUMENT_FOLDER_CREATE',
                f'Created folder: {folder.name}',
                request
            )

            messages.success(request, 'Folder created successfully!')
            return redirect('documents:document_list')
    else:
        form = DocumentFolderForm()

    return render(request, 'documents/folder_form.html', {
        'form': form,
        'title': 'New Folder'
    })


@login_required
@manager_or_admin_required
def folder_update(request, pk):
    """Update a document folder"""
    folder = get_object_or_404(DocumentFolder, pk=pk)

    if request.method == 'POST':
        form = DocumentFolderForm(request.POST, instance=folder)
        if form.is_valid():
            folder = form.save()

            log_audit(
                request.user,
                'DOCUMENT_FOLDER_UPDATE',
                f'Updated folder: {folder.name}',
                request
            )

            messages.success(request, 'Folder updated successfully!')
            return redirect('documents:document_list')
    else:
        form = DocumentFolderForm(instance=folder)

    return render(request, 'documents/folder_form.html', {
        'form': form,
        'title': 'Rename Folder',
        'folder': folder
    })


@login_required
def document_upload(request):
    """Upload a new document"""
    if request.method == 'POST':
        form = DocumentUploadForm(request.POST, request.FILES)
        if form.is_valid():
            document = form.save(commit=False)
            document.owner = request.user
            uploaded_file = form.cleaned_data.get('file')
            if uploaded_file:
                document.file_size = uploaded_file.size
                document.file_type = uploaded_file.content_type
            else:
                document.file_size = 0
                if document.google_docs_url:
                    document.file_type = 'Google Docs'
                elif document.google_sheets_url:
                    document.file_type = 'Google Sheets'
            document.save()
            
            log_audit(
                request.user,
                'DOCUMENT_UPLOAD',
                f'Uploaded document: {document.title}',
                request
            )
            
            messages.success(request, 'Document uploaded successfully!')
            return redirect('documents:document_detail', pk=document.pk)
    else:
        form = DocumentUploadForm()
    
    return render(request, 'documents/document_upload.html', {'form': form})


@login_required
def document_detail(request, pk):
    """View document details"""
    document = get_object_or_404(Document, pk=pk)
    
    # Check if document is archived
    if document.is_archived:
        messages.error(request, 'This document has been archived and is no longer available.')
        return redirect('documents:document_list')
    
    # Check access permission
    if not can_access_document(request.user, document):
        messages.error(request, 'You do not have permission to view this document.')
        return redirect('documents:document_list')
    
    log_audit(
        request.user,
        'DOCUMENT_VIEW',
        f'Viewed document: {document.title}',
        request
    )

    preview_type = None
    preview_context = {
        'preview_text': '',
        'preview_rows': [],
        'preview_sheet_name': '',
        'preview_truncated': False,
        'preview_error': '',
        'preview_char_limit': PREVIEW_CHAR_LIMIT,
        'preview_row_limit': PREVIEW_ROW_LIMIT,
        'preview_column_limit': PREVIEW_COLUMN_LIMIT,
    }

    if document.file:
        file_extension = document.get_file_extension()
        if os.path.exists(document.file.path):
            if not _is_safe_media_path(document.file.path):
                preview_type = 'unsupported'
                preview_context['preview_error'] = 'Preview is unavailable for this file.'
                return render(request, 'documents/document_detail.html', {
                    'document': document,
                    'preview_type': preview_type,
                    **preview_context
                })
            file_size = document.file_size or document.file.size
            if file_size > PREVIEW_MAX_FILE_SIZE:
                preview_type = 'unsupported'
                preview_context['preview_error'] = (
                    'Preview is available for files up to 5 MB. Please download to view.'
                )
                return render(request, 'documents/document_detail.html', {
                    'document': document,
                    'preview_type': preview_type,
                    **preview_context
                })
            try:
                if file_extension == '.pdf':
                    preview_type = 'pdf'
                elif file_extension in ('.jpg', '.jpeg', '.png'):
                    preview_type = 'image'
                elif file_extension in ('.txt', '.csv', '.log'):
                    preview_type = 'text'
                    preview_text, preview_truncated = _load_text_preview(document.file.path)
                    preview_context.update({
                        'preview_text': preview_text,
                        'preview_truncated': preview_truncated,
                    })
                elif file_extension == '.docx':
                    preview_type = 'text'
                    preview_text, preview_truncated = _load_docx_preview(document.file.path)
                    preview_context.update({
                        'preview_text': preview_text,
                        'preview_truncated': preview_truncated,
                    })
                elif file_extension == '.xlsx':
                    preview_type = 'spreadsheet'
                    sheet_name, preview_rows, preview_truncated = _load_spreadsheet_preview(
                        document.file.path
                    )
                    preview_context.update({
                        'preview_sheet_name': sheet_name,
                        'preview_rows': preview_rows,
                        'preview_truncated': preview_truncated,
                    })
                elif file_extension in ('.doc', '.xls'):
                    preview_type = 'office'
                else:
                    preview_type = 'unsupported'
            except PreviewError:
                preview_type = 'unsupported'
                preview_context['preview_error'] = (
                    'Preview could not be generated because the file contents could not be read.'
                )
            except OSError:
                preview_type = 'unsupported'
                preview_context['preview_error'] = (
                    'Preview could not be generated because the file could not be accessed.'
                )
        else:
            preview_context['preview_error'] = 'Document file not found for preview.'
            preview_type = 'unsupported'
    elif document.google_docs_url:
        preview_type = 'google_docs'
    elif document.google_sheets_url:
        preview_type = 'google_sheets'
    else:
        preview_type = 'unsupported'

    return render(request, 'documents/document_detail.html', {
        'document': document,
        'preview_type': preview_type,
        **preview_context
    })


@login_required
@xframe_options_sameorigin
def document_preview(request, pk):
    """Preview a document inline"""
    document = get_object_or_404(Document, pk=pk)

    if document.is_archived:
        messages.error(request, 'This document has been archived and is no longer available.')
        return redirect('documents:document_list')

    if not can_access_document(request.user, document):
        messages.error(request, 'You do not have permission to view this document.')
        return redirect('documents:document_list')

    if not document.file:
        messages.info(
            request,
            'Preview is not available for documents linked to external services like Google Docs.'
        )
        return redirect('documents:document_detail', pk=pk)

    file_extension = document.get_file_extension()
    if file_extension not in ('.pdf', '.jpg', '.jpeg', '.png'):
        messages.info(request, 'Inline preview is not available for this file type.')
        return redirect('documents:document_detail', pk=pk)

    if os.path.exists(document.file.path):
        response = FileResponse(document.file.open('rb'), content_type=document.file_type)
        filename = quote(os.path.basename(document.file.name))
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response

    raise Http404("Document file not found")


@login_required
def document_download(request, pk):
    """Download a document"""
    document = get_object_or_404(Document, pk=pk)
    
    # Check if document is archived
    if document.is_archived:
        messages.error(request, 'This document has been archived and cannot be downloaded.')
        return redirect('documents:document_list')
    
    # Check access permission
    if not can_access_document(request.user, document):
        messages.error(request, 'You do not have permission to download this document.')
        return redirect('documents:document_list')

    if not document.file:
        messages.error(request, 'This document does not have a file to download.')
        return redirect('documents:document_detail', pk=pk)
    
    log_audit(
        request.user,
        'DOCUMENT_DOWNLOAD',
        f'Downloaded document: {document.title}',
        request
    )
    
    # Serve file
    if os.path.exists(document.file.path):
        response = FileResponse(document.file.open('rb'), as_attachment=True)
        return response
    else:
        raise Http404("Document file not found")


@login_required
def document_update(request, pk):
    """Update document metadata"""
    document = get_object_or_404(Document, pk=pk)
    
    # Check if document is archived
    if document.is_archived:
        messages.error(request, 'This document has been archived and cannot be updated.')
        return redirect('documents:document_list')
    
    # Only owner, manager, or admin can update
    if not (document.owner == request.user or request.user.is_president or request.user.is_adviser):
        messages.error(request, 'You do not have permission to update this document.')
        return redirect('documents:document_detail', pk=pk)
    
    if request.method == 'POST':
        form = DocumentUpdateForm(request.POST, instance=document)
        if form.is_valid():
            form.save()
            
            log_audit(
                request.user,
                'DOCUMENT_UPDATE',
                f'Updated document: {document.title}',
                request
            )
            
            messages.success(request, 'Document updated successfully!')
            return redirect('documents:document_detail', pk=pk)
    else:
        form = DocumentUpdateForm(instance=document)
    
    return render(request, 'documents/document_update.html', {
        'form': form,
        'document': document
    })


@login_required
@manager_or_admin_required
def document_delete(request, pk):
    """Archive a document (soft delete)"""
    document = get_object_or_404(Document, pk=pk)
    
    # Only owner, manager, or admin can archive
    if not (document.owner == request.user or request.user.is_president or request.user.is_adviser):
        messages.error(request, 'You do not have permission to archive this document.')
        return redirect('documents:document_detail', pk=pk)
    
    if request.method == 'POST':
        title = document.title
        
        # Archive the document instead of deleting
        document.is_archived = True
        document.archived_at = timezone.now()
        document.archived_by = request.user
        document.save()
        
        log_audit(
            request.user,
            'DOCUMENT_ARCHIVE',
            f'Archived document: {title}',
            request
        )
        
        messages.success(request, 'Document archived successfully!')
        return redirect('documents:document_list')
    
    return render(request, 'documents/document_delete.html', {'document': document})
